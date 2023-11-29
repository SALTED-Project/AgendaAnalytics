#!/usr/local/bin/python3

import configparser
from datetime import datetime


config = configparser.ConfigParser()
configFilePath = 'org_to_excel.ini'
config.read(configFilePath)

server_url  =config.get('EXCEL', 'server_url')
score_threshold = config.getfloat('EXCEL', 'score_threshold')
template_file = config.get('INPUT','template_path')
sdg_template_file = config.get('INPUT','sdg_template_path')
report_path=config.get('OUTPUT', 'report_path')

scorpio_api = config.get('SCORPIO', 'api')
scorpio_api_pub = config.get('SCORPIO', 'api_pub')
fileserver_api = config.get('FILESERVER', 'api')
context_organization = config.get('SCORPIO', 'context_organization')
context_dsr = config.get('SCORPIO', 'context_dsr')
context_agenda = config.get('SCORPIO', 'context_agenda')
context_kpi = config.get('SCORPIO', 'context_kpi')


print('Content-Type: text/html \n\n')
print('')

# insert the necessary libraries
from openpyxl import load_workbook #pip install openpyxl


import cgi    
import cgitb
from datetime import datetime 
import requests
import numpy as np
import os, sys
import collections


# help for debugging
# disable
def blockPrint():
    sys.stdout = open(os.devnull, 'w')
# restore
def enablePrint():
    sys.stdout = sys.__stdout__



# get supplied parameters within URL
cgitb.enable()
cgitb.enable(display=1,logdir=None,context=5,format='html')
form = cgi.FieldStorage()

if ("org_entity_id" not in form) or ("kpi_entity_url" not in form) or ("agenda_entity_url" not in form)  :
    print('Name of the company or corresponding matching entity / agenda entity is not available. Are you sure that the crawling & matching has already taken place for this organization?')
else:  

    # disable print statements fro production
    blockPrint()   
    
    def get_entities_by_id(id, context):
        try:
            # access entities by type within the scorpio broker, using specific context links provided in header
            url = scorpio_api+"/"+id     
            headers = {'Accept': 'application/ld+json','Link': f'<{context}>; rel="http://www.w3.org/ns/json-ld#context"; type="application/ld+json"'}
            r = requests.get(url, headers=headers) 
            print("Response from Scorpio for {id}: {status_code}".format(id=id,status_code=r.status_code))          
            if r.status_code==200:
                json_entity=r.json()
            else:
                json_entity=None
        except:
            json_entity= None
        return json_entity
          
        

    
    try:
        org_entity_id = form["org_entity_id"].value
        
        kpi_entity_url = form["kpi_entity_url"].value
        kpi_entity_id = kpi_entity_url.split("/")[-1]
        
        agenda_entity_url = form["agenda_entity_url"].value
        agenda_entity_id = agenda_entity_url.split("/")[-1]
        
        # load organization
        print("getting organization entity of interest")
        org=get_entities_by_id(org_entity_id, context_organization)
        org_name=org["name"]["value"]        

        # load agenda
        print("getting agenda entity of interest")
        agenda=get_entities_by_id(agenda_entity_id, context_agenda)
        agenda_name=agenda["name"]["value"]         

        # load kpi
        print("getting kpi entity of interest")
        kpi=get_entities_by_id(kpi_entity_id, context_kpi)     
        fileserver_url = fileserver_api+"/"+(kpi["kpiValue"]["value"]["complete_values"]).split("/files/")[-1]
        r = requests.get(fileserver_url)
        kpi_value_full = r.json()      
        
        mean_per_level1_detailed_raw = kpi_value_full["detailed"]["mean_per_level1"]    # example: {'level0': '1', 'level1': '3', 'similarity': 0.2878727668858304}
        mean_per_level0_detailed_raw= kpi_value_full["detailed"]["mean_per_level0"]
        values_detailed_raw = kpi_value_full["detailed"]["raw_values"]  # example: {'level0': '2', 'level1': '4', 'fragment': 29.0, 'similarity': 0.4685954451560974} 
        agenda_info_raw = kpi_value_full["text"]["reference"] 
        list_level0 = [value["level0"] for value in mean_per_level0_detailed_raw]       

        # adapt regarding threshold
        mean_per_level1_detailed = []
        for item in mean_per_level1_detailed_raw:
            level0 = item["level0"]
            level1 = item["level1"]
            similarity_values = []
            for raw_value in values_detailed_raw: 
                if raw_value["level0"]==level0 and raw_value["level1"]==level1:
                    similarity_values.append(raw_value["similarity"])
            # only analysis sentences with score above threshold form the target score
            mean_level1 = np.nanmean( [similarity if similarity > score_threshold else np.nan for similarity in similarity_values ])
            mean_per_level1_detailed.append({'level0': level0, 'level1': level1, 'similarity': mean_level1 })

        sim_per_level0_detailed ={}
        mean_per_level0_detailed={}           
        for x in list_level0:
            sim_per_level0_detailed[x]=[]
        for value in mean_per_level1_detailed:
            level0 = value["level0"]
            sim_value = value["similarity"] # only when all values were below the target value is np.nan
            sim_per_level0_detailed[level0].append(sim_value)          
        for x in list_level0:             
            mean_value_detailed = np.nansum(sim_per_level0_detailed[x])/len(sim_per_level0_detailed[x]) # form the goal mean from all targets, even when they are np.nan               
            mean_per_level0_detailed[x] = mean_value_detailed

        
        
        # set empty SDG template
        if agenda_entity_id=="insert full sdg agenda urn here":
           template_file=sdg_template_file
           wb = load_workbook(template_file) # load the template
           agenda_template_name="Template_{agenda}.xlsx".format(agenda = agenda_entity_id)
           wb.save(report_path+agenda_template_name) 
           print("SDG template used")
        
        else:
            agenda_template_name="Template_{agenda}.xlsx".format(agenda = agenda_entity_id)
            # check if template is precalculated
            if os.path.exists(report_path+agenda_template_name):
                print("agenda specific template exists")
            else:
                print("agenda specific template will be generated")
                # create specific agenda template             
                template_file=template_file
                agenda_files=collections.defaultdict(dict)           
                for item in agenda_info_raw:              
                    # create list of goal title and target-title / text
                    agenda_files[item["level0"]][item["level1"]]=[item["title"], item["text"]]                 

                agenda_files = collections.OrderedDict(sorted(agenda_files.items()))
                print(agenda_files)
                            
                wb = load_workbook(template_file, data_only=False, rich_text=True) # load the template
                sheets = wb.sheetnames # identify the sheets in the template
                
                goals_index = sheets.index("Goals")
                targets_index=sheets.index("Targets")      
                analyse_index = sheets.index("Analyse")
                reftargets_index=sheets.index("Ref_Targets")
                refgoals_index=sheets.index("Ref_Goals")
                
                Sheet_Goals = wb[sheets[goals_index]] 
                Sheet_RefGoals = wb[sheets[refgoals_index]] 
                Sheet_Analyse = wb[sheets[analyse_index]] 
                Sheet_Targets = wb[sheets[targets_index]] 
                Sheet_RefTargets = wb[sheets[reftargets_index]]      
                wb.active=0 

                # set agenda on sheet headers
                Sheet_Goals.cell(row = 3, column = 2).value = agenda_name
                Sheet_Targets.cell(row = 3, column = 2).value = agenda_name
                Sheet_Analyse.cell(row = 3, column = 2).value = agenda_name


                # fill agenda info
                i=0
                j=0
                for level0, level1_dict in dict(agenda_files).items():
                    Sheet_Goals.cell(row = 6+i, column = 3).value = level0
                    Sheet_RefGoals.cell(row = 2+i, column = 1).value = level0
                    # pick first example = target1, because it always exists, to fill goal title
                    Sheet_Goals.cell(row = 6+i, column = 4).value = level1_dict["1"][0]
                    for level1,content in level1_dict.items():
                        print(f"{level0}-{level1}")
                        Sheet_Targets.cell(row = 6+j, column = 3).value = level0  
                        Sheet_Targets.cell(row = 6+j, column = 5).value = level1 
                        Sheet_Targets.cell(row = 6+j, column = 6).value = content[1]
                        Sheet_RefTargets.cell(row = 2+j, column = 1).value = level0  
                        Sheet_RefTargets.cell(row = 2+j, column = 2).value = level1 
                        Sheet_RefTargets.cell(row = 2+j, column = 6).value =  content[1]
                        j=j+1
                    i = i+1    
                wb.save(report_path+agenda_template_name) 
                print("saved agenda specific template")



        # set filename & path for company excel
        file_name="Compliance_Report_{agenda}_{co}.xlsx".format(co=org_entity_id, agenda = agenda_entity_id)
        file_name_ts="Compliance_Report_{agenda}_{co}_{ts}.xlsx".format(co=org_entity_id, agenda = agenda_entity_id, ts=datetime.now().strftime("%Y%m%d-%H%M%S-%f"))
              
        
        
        # create company excel from agenda template 
        wb = load_workbook(report_path+agenda_template_name, data_only=False) # load the template
        sheets = wb.sheetnames # identify the sheets in the template
        
        goals_index = sheets.index("Goals")
        targets_index=sheets.index("Targets")
        analyse_index=sheets.index("Analyse")        
        reftargets_index=sheets.index("Ref_Targets")  
        refgoals_index=sheets.index("Ref_Goals")  
        document_index=sheets.index("Document")  
        process_index=sheets.index("Process_Input")  
        
        Sheet_Goals = wb[sheets[goals_index]]
        Sheet_Goals.cell(row = 2, column = 2).value = org_name
        Sheet_Targets = wb[sheets[targets_index]] 
        Sheet_Targets.cell(row = 2, column = 2).value = org_name
        Sheet_Analyse = wb[sheets[analyse_index]] 
        Sheet_Analyse.cell(row = 2, column = 2).value = org_name
        Sheet_RefTargets = wb[sheets[reftargets_index]]   
        Sheet_RefGoals = wb[sheets[refgoals_index]]   
        Sheet_Document = wb[sheets[document_index]] 
        Sheet_Process = wb[sheets[process_index]] 
        wb.active=0 
        

        print(mean_per_level0_detailed)
        print(mean_per_level1_detailed)


        print("fill RefGoals") 
        # fill sheet Ref_Goals (show mean of targets above threshold)   
        for rowNum in range(2, len(mean_per_level0_detailed)+2):
            level0 = str(Sheet_RefGoals.cell(row=rowNum, column=1).value)
            Sheet_RefGoals.cell(row=rowNum, column=3).value = mean_per_level0_detailed[level0]

    
        print("fill RefTargets")           
        # fill sheet Ref_Targets (shows targets above threshold, otherwise np.nan)
        for rowNum in range(2, len(mean_per_level1_detailed)+2):
            level0 = str(Sheet_RefTargets.cell(row=rowNum, column=1).value)
            level1 = str(Sheet_RefTargets.cell(row=rowNum, column=2).value)
            Sheet_RefTargets.cell(row=rowNum, column=7).value = next(filter(lambda item: item["level0"]==level0 and item["level1"]==level1 , mean_per_level1_detailed), None)["similarity"]   
        
        print("fill Document") 
        # fill sheet Document      
        for item in kpi_value_full["text"]["analysis"]:
            prepared_row_document= [item["fragment"], item["text"]]
            Sheet_Document.append(prepared_row_document)
        
        print("fill ProcessInput") 
        # fill sheet Process_Input (shows raw values for transparency, since all where used to calculate tragets)         
        for item in values_detailed_raw:
            level0_title = next(filter(lambda agenda_item: agenda_item["level0"]==item["level0"], agenda_info_raw), None)["title"]
            prepared_row = [level0_title+"."+item["level0"]+"-"+item["level1"]+".txt", item["level0"]+"-"+item["level1"], item["level0"],item["level1"],item["similarity"], item["fragment"]]
            Sheet_Process.append(prepared_row)           

        
        # save final excel files        
        wb.save(report_path+file_name) #save the new excel in the file_path under the generated name
        wb.save(report_path+file_name_ts) #save the new excel in the file_path under the generated name





        print("final excel saved at:")
        print(report_path+file_name)
        
        enablePrint()
        # open excel in browser (server_url must be reachable from public, otherwise error message is shown in browser window)
        print("<html>")
    
        print("<head>")
        print("<title>Welcome to Agenda Analytics</title>")
        print("</head>")

        print("<body>")

        print("<iframe src= "+ "https://view.officeapps.live.com/op/embed.aspx?src={url}Compliance_Report_{agenda}_{co}.xlsx".format(url=server_url,co=org_entity_id, agenda=agenda_entity_id) +" width='100%' height='825px' frameborder='0' rel='noopener noreferrer' ></iframe>")

        print("</body>")
        print("</html>")

    except Exception as e:
        print("An error occured while generating the excel. Empty template may be shown. Error: {}".format(e))



    
        
    
    


